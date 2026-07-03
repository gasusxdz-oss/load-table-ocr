from functools import wraps

from flask import Flask
from flask import render_template
from flask import request
from flask import jsonify
from flask import send_file
from flask import redirect
from flask import url_for
from flask import session

import fitz
import io
import json
import os
import tempfile
import requests
from rapidfuzz import process
from rapidfuzz import fuzz

from dotenv import load_dotenv


load_dotenv()

app = Flask(__name__, static_folder="static", template_folder="templates")
app.config["SECRET_KEY"] = os.environ.get("FLASK_SECRET_KEY") or os.urandom(24)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
TEMP_FOLDER = os.path.join(BASE_DIR, "tmp")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(TEMP_FOLDER, exist_ok=True)


# ==================================================
# RapidFuzz補正設定
# ==================================================

AUTO_THRESHOLD = 80
CHECK_THRESHOLD = 60

TEXT_REPLACE = {
    "一": "-",
    "ー": "-",
    "s9": "sq",
    "S9": "sq",
    "(1CE)": "(1C-E)",
    "(1C E)": "(1C-E)"
}

# Firebase設定
FIREBASE_SERVICE_ACCOUNT_PATH = os.getenv(
    "FIREBASE_SERVICE_ACCOUNT_PATH",
    os.path.join(BASE_DIR, "load-table-ocr-firebase-adminsdk-fbsvc-03a4b912a0.json")
).strip()
FIREBASE_CONFIG_PATH = os.getenv(
    "FIREBASE_CONFIG_PATH",
    os.path.join(BASE_DIR, "firebaseConfig.json")
).strip()
FIREBASE_COLLECTION = os.getenv("FIREBASE_EQUIPMENT_COLLECTION", "equipment_dictionary").strip()
FIREBASE_DOCUMENT = os.getenv("FIREBASE_EQUIPMENT_DOCUMENT", "main").strip()

FIREBASE_API_KEY = os.getenv("FIREBASE_API_KEY", "").strip()
FIREBASE_PROJECT_ID = os.getenv("FIREBASE_PROJECT_ID", "").strip()
FIREBASE_AUTH_URL = ""

if FIREBASE_API_KEY:
    FIREBASE_AUTH_URL = (
        f"https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword?key={FIREBASE_API_KEY}"
    )

if not FIREBASE_API_KEY and os.path.exists(FIREBASE_CONFIG_PATH):
    try:
        with open(FIREBASE_CONFIG_PATH, "r", encoding="utf-8") as f:
            fb_conf = json.load(f)
            FIREBASE_API_KEY = str(fb_conf.get("apiKey", "")).strip()
            FIREBASE_PROJECT_ID = str(fb_conf.get("projectId", "")).strip()
            if FIREBASE_API_KEY:
                FIREBASE_AUTH_URL = (
                    f"https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword?key={FIREBASE_API_KEY}"
                )
    except Exception as e:
        print(f"firebaseConfig.json の読み込みエラー: {e}")
elif not FIREBASE_API_KEY:
    print(f"firebaseConfig.json が見つかりません: {FIREBASE_CONFIG_PATH}")

# 設備辞書の読み込み
EQUIPMENT_LIST = []
OCR_DICT = {}
DICT_LOADED = False


def firebase_sign_in(email: str, password: str):
    """Firebase Auth REST API でログインする"""
    if not FIREBASE_API_KEY:
        raise RuntimeError("FIREBASE_API_KEY が未設定です")

    payload = {
        "email": email,
        "password": password,
        "returnSecureToken": True,
    }

    response = requests.post(FIREBASE_AUTH_URL, json=payload, timeout=10)
    if response.status_code == 200:
        return response.json()

    raise ValueError(response.json().get("error", {}).get("message", "Firebase認証に失敗しました"))


def load_equipment_dict():
    global EQUIPMENT_LIST, OCR_DICT, DICT_LOADED

    EQUIPMENT_LIST = []
    OCR_DICT = {}
    DICT_LOADED = False

    if not FIREBASE_SERVICE_ACCOUNT_PATH or not os.path.exists(FIREBASE_SERVICE_ACCOUNT_PATH):
        print("Firebaseサービスアカウント情報が未設定のため、辞書を読み込みません")
        return

    try:
        import firebase_admin
        from firebase_admin import credentials, firestore

        if not firebase_admin._apps:
            cred = credentials.Certificate(FIREBASE_SERVICE_ACCOUNT_PATH)
            firebase_admin.initialize_app(cred)

        db = firestore.client()
        doc = db.collection(FIREBASE_COLLECTION).document(FIREBASE_DOCUMENT).get()

        if not doc.exists:
            print("Firestoreに設備辞書が存在しません")
            return

        data = doc.to_dict() or {}

        equipment_names = data.get("equipment_names")
        if isinstance(equipment_names, list):
            EQUIPMENT_LIST = [str(item).strip() for item in equipment_names if str(item).strip()]
        elif isinstance(data.get("equipment"), list):
            EQUIPMENT_LIST = [str(item).strip() for item in data.get("equipment") if str(item).strip()]

        raw_ocr_pairs = data.get("ocr_corrections")
        if isinstance(raw_ocr_pairs, dict):
            OCR_DICT = {str(k).strip(): str(v).strip() for k, v in raw_ocr_pairs.items() if str(k).strip()}
        elif isinstance(raw_ocr_pairs, list):
            for item in raw_ocr_pairs:
                if isinstance(item, dict):
                    wrong = str(item.get("wrong") or item.get("from") or item.get("source") or "").strip()
                    correct = str(item.get("correct") or item.get("to") or item.get("target") or "").strip()
                    if wrong and correct:
                        OCR_DICT[wrong] = correct

        DICT_LOADED = bool(EQUIPMENT_LIST or OCR_DICT)
        print(f"Firebaseから設備辞書を読み込みました: {len(EQUIPMENT_LIST)}項目")
    except Exception as e:
        print(f"Firebase設備辞書の読み込みエラー: {e}")


# easyocr リーダーの遅延初期化（メモリ節約）
reader = None

def get_reader():
    global reader
    if reader is None:
        import easyocr

        reader = easyocr.Reader(
            ['ja', 'en'],
            gpu=False,
            model_storage_directory=os.path.join(BASE_DIR, 'models'),
            user_network_directory=os.path.join(BASE_DIR, 'models')
        )
    return reader

current_image_bytes = None
current_image_path = None
current_excel_bytes = None
current_excel_path = None
current_pdf_path = None
current_pdf_page_count = 0
current_pdf_page_number = 0
progress = 0


def login_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if not session.get("usr"):
            if request.is_json:
                return jsonify({"error": "ログインが必要です"}), 401
            return redirect(url_for("login"))
        return view_func(*args, **kwargs)

    return wrapped


def normalize(text):
    """テキストの正規化"""
    if text is None:
        return ""

    try:
        import math
        if isinstance(text, float) and math.isnan(text):
            return ""
    except Exception:
        pass
    
    text = str(text)
    
    for before, after in TEXT_REPLACE.items():
        text = text.replace(before, after)
    
    return text.strip()


def split_equipment_name(text):
    """
    設備名称を基本名称と補足情報に分割
    例: 排気ファソ(空調機械室1-2) → (排気ファソ, (空調機械室1-2))
    """
    text = str(text).strip()
    
    paren_pos = len(text)
    
    for mark in ["(", "（"]:
        pos = text.find(mark)
        if pos >= 0:
            paren_pos = min(paren_pos, pos)
    
    if paren_pos < len(text):
        return (text[:paren_pos].strip(), text[paren_pos:])
    
    return text, ""


def correct_equipment_name(original):
    """
    RapidFuzzを使ってOCR結果を補正
    戻り値: (補正結果, 類似度, 判定)
    """
    
    if not DICT_LOADED:
        load_equipment_dict()
        if not DICT_LOADED:
            return original, 0, "未補正"
    
    original = str(original).strip()
    corrected = original
    score = 0
    status = ""
    
    # OCR辞書の完全一致
    if original in OCR_DICT:
        corrected = OCR_DICT[original]
        score = 100
        status = "OCR辞書"
        return corrected, score, status
    
    # 設備名部分だけ抽出
    base_name, suffix = split_equipment_name(original)
    
    # OCR辞書で部分補正
    temp_name = base_name
    
    for wrong, right in sorted(
        OCR_DICT.items(),
        key=lambda x: len(x[0]),
        reverse=True
    ):
        if temp_name == wrong:
            temp_name = right
            break
    
    # RapidFuzzで類似度を計算
    if EQUIPMENT_LIST:
        match = process.extractOne(
            temp_name,
            EQUIPMENT_LIST,
            scorer=fuzz.WRatio
        )
        
        if match:
            candidate = match[0]
            score = round(match[1], 1)
            
            if score >= 95:
                corrected = candidate + suffix
                status = "RapidFuzz自動補正"
            elif score >= 85:
                corrected = candidate + suffix
                status = "要確認"
            else:
                corrected = original
                status = "未補正"
        else:
            corrected = original
            status = "未補正"
    else:
        corrected = original
        status = "未補正"
    
    return corrected, score, status


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template("login.html", msg=None, firebase_ready=bool(FIREBASE_API_KEY))

    email = request.form.get("email", "").strip()
    password = request.form.get("password", "")

    if not email or not password:
        return render_template("login.html", msg="メールアドレスとパスワードを入力してください。", firebase_ready=bool(FIREBASE_API_KEY))

    if not FIREBASE_API_KEY:
        return render_template("login.html", msg="Firebaseの設定が未完了です。環境変数 FIREBASE_API_KEY を設定してください。", firebase_ready=False)

    try:
        result = firebase_sign_in(email, password)
        session["usr"] = result.get("email") or email
        session["id_token"] = result.get("idToken")
        return redirect(url_for("index"))
    except Exception as exc:
        return render_template("login.html", msg=str(exc), firebase_ready=bool(FIREBASE_API_KEY))


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
def index():
    if not session.get("usr"):
        return redirect(url_for("login"))

    return render_template(
        "index.html",
        usr=session.get("usr")
    )


@app.route(
    "/upload",
    methods=["POST"]
)
@login_required
def upload():

    global current_image_bytes
    global current_image_path
    global current_pdf_path
    global current_pdf_page_count
    global current_pdf_page_number

    file = request.files["pdf"]

    pdf_path = os.path.join(
        UPLOAD_FOLDER,
        file.filename
    )

    file.save(pdf_path)

    with fitz.open(pdf_path) as doc:
        current_pdf_path = pdf_path
        current_pdf_page_count = len(doc)
        current_pdf_page_number = 0

        page = doc[0]
        pix = page.get_pixmap(
            matrix=fitz.Matrix(3, 3)
        )

        with tempfile.NamedTemporaryFile(suffix=".png", dir=TEMP_FOLDER, delete=False) as tmp:
            image_path = tmp.name

        pix.save(image_path)
        current_image_bytes = pix.tobytes("png")
        current_image_path = image_path
        session["image_path"] = image_path

    return jsonify({
        "image": "/page.png",
        "total_pages": current_pdf_page_count
    })


@app.route(
    "/page",
    methods=["POST"]
)
@login_required
def change_page():

    global current_image_bytes
    global current_image_path
    global current_pdf_path
    global current_pdf_page_number
    global current_pdf_page_count

    if not current_pdf_path or not os.path.exists(current_pdf_path):
        return jsonify({
            "error": "PDFが読み込まれていません"
        }), 400

    data = request.json
    page_number = int(data["page"])

    if page_number < 0 or page_number >= current_pdf_page_count:
        return jsonify({
            "error": "ページ番号が無効です"
        }), 400

    current_pdf_page_number = page_number

    with fitz.open(current_pdf_path) as doc:
        page = doc[page_number]
        pix = page.get_pixmap(
            matrix=fitz.Matrix(3, 3)
        )

        with tempfile.NamedTemporaryFile(suffix=".png", dir=TEMP_FOLDER, delete=False) as tmp:
            image_path = tmp.name

        pix.save(image_path)
        current_image_bytes = pix.tobytes("png")
        current_image_path = image_path
        session["image_path"] = image_path

    return jsonify({
        "image": "/page.png",
        "current_page": current_pdf_page_number + 1,
        "total_pages": current_pdf_page_count
    })


@app.route(
    "/save_selection",
    methods=["POST"]
)
@login_required
def save_selection():

    global current_image_bytes
    global current_excel_bytes
    global current_excel_path

    from PIL import Image
    import numpy as np
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    data = request.json

    left = int(data["left"])
    top = int(data["top"])
    right = int(data["right"])
    bottom = int(data["bottom"])

    rows = int(data["rows"])

    if not current_image_bytes:
        return jsonify({"error": "画像データがありません"}), 400

    img = Image.open(io.BytesIO(current_image_bytes))

    width = right - left
    height = bottom - top

    row_height = height / rows

    wb = Workbook()
    
    # OCR結果シート
    ws_ocr = wb.active
    ws_ocr.title = "OCR結果"
    
    ws_ocr["A1"] = "OCR結果"
    
    # 補正結果シート
    ws_corrected = wb.create_sheet("補正結果")
    ws_corrected["A1"] = "行番号"
    ws_corrected["B1"] = "OCR結果"
    ws_corrected["C1"] = "補正結果"
    ws_corrected["D1"] = "類似度"
    ws_corrected["E1"] = "判定"

    global progress
    progress = 0

    for i in range(rows):

        progress = int(
            (i + 1) / rows * 100
        )

        y1 = int(
            top +
            row_height * i
        )

        y2 = int(
            top +
            row_height * (i + 1)
        )

        crop = img.crop(
            (
                left,
                y1,
                right,
                y2
            )
        )

        crop_np = np.asarray(crop)

        ocr_reader = get_reader()
        result = ocr_reader.readtext(
            crop_np,
            detail=0
        )

        text = " ".join(result)
        
        # OCR結果シートに出力
        ws_ocr.cell(
            row=i + 2,
            column=1
        ).value = text

        # 補正処理
        corrected, score, status = correct_equipment_name(text)
        
        # 補正結果シートに出力
        ws_corrected.cell(row=i + 2, column=1).value = i + 1
        ws_corrected.cell(row=i + 2, column=2).value = text
        ws_corrected.cell(row=i + 2, column=3).value = corrected
        ws_corrected.cell(row=i + 2, column=4).value = score
        ws_corrected.cell(row=i + 2, column=5).value = status
        
        # 補正された場合は色付け
        if text != corrected:
            yellow = PatternFill(
                fill_type="solid",
                fgColor="FFFF00"
            )
            ws_corrected.cell(row=i + 2, column=2).fill = yellow
            ws_corrected.cell(row=i + 2, column=3).fill = yellow

        print(
            f"{i+1}: {text} → {corrected} ({status})"
        )

    with tempfile.NamedTemporaryFile(suffix=".xlsx", dir=TEMP_FOLDER, delete=False) as tmp:
        excel_path = tmp.name

    wb.save(excel_path)
    current_excel_path = excel_path
    session["excel_path"] = excel_path

    return jsonify({
        "status": "ok",
        "excel": "/download_excel"
    })

@app.route("/progress")
@login_required
def get_progress():

    return jsonify({
        "progress": progress
    })

@app.route("/page.png")
@login_required
def serve_page_image():
    image_path = session.get("image_path") or current_image_path
    if not image_path or not os.path.exists(image_path):
        return jsonify({"error": "画像データがありません"}), 404

    return send_file(
        image_path,
        mimetype="image/png",
        download_name="page.png"
    )


@app.route(
    "/download_excel"
)
@login_required
def download_excel():
    excel_path = session.get("excel_path") or current_excel_path
    if not excel_path or not os.path.exists(excel_path):
        return jsonify({"error": "Excelデータがありません"}), 404

    return send_file(
        excel_path,
        as_attachment=True,
        download_name="ocr_result.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(
        host="0.0.0.0",
        port=port,
        debug=False
    )